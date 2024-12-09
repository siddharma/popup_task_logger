import tkinter as tk
from tkinter import simpledialog, messagebox
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import time

# Excel file setup
EXCEL_FILE = "task_log.xlsx"

def setup_excel():
    """Initialize the Excel file with headers if it doesn't exist."""
    if not os.path.exists(EXCEL_FILE):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Task Log"
        # Set up headers (Days and Times)
        sheet.append(["Date"] + [f"{hour:02d}:{minute:02d}" 
                                 for hour in range(24) for minute in [0, 30]])
        workbook.save(EXCEL_FILE)

def save_task_to_excel(task, timestamp):
    """Save the task into the correct cell based on time."""
    date_str = timestamp.strftime("%Y-%m-%d")
    time_str = timestamp.strftime("%H:%M")
    
    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active
    
    # Find the row for the date or create one
    date_row = None
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == date_str:
            date_row = row
            break
    if not date_row:
        date_row = sheet.max_row + 1
        sheet.cell(row=date_row, column=1, value=date_str)
    
    # Find the correct column for the time
    time_col = None
    for col in range(2, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == time_str or (sheet.cell(row=1, column=col).value > time_str):
            time_col = col
            break
    
    # Save the task in the appropriate cell
    if time_col:
        #sheet.cell(row=date_row, column=time_col, value=task)
        ##old_value = sheet.cell(row=date_row, column=time_col).value;
        sheet.cell(row=date_row, column=time_col, value=task)
    
    workbook.save(EXCEL_FILE)

def show_popup():
    """Display a popup form to collect task information."""
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    
    task = simpledialog.askstring("Task Entry", "Enter your task:", parent=root)
    
    if task:
        timestamp = datetime.now()
        save_task_to_excel(task, timestamp)
        messagebox.showinfo("Task Saved", "Your task has been logged successfully!")
    root.destroy()

def main():
    """Main function to run the periodic task popup."""
    setup_excel()
    
    while True:
        show_popup()
        time.sleep(60*30)  # Wait for 30 minutes (30 * 60 seconds)

if __name__ == "__main__":
    main()